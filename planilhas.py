"""
planilhas.py — Gera as planilhas Excel do robo.

Duas coisas diferentes:

1. Planilha de notas, por empresa e competencia (vai na raiz RELATORIOS):
   `nfses_emitidas_{CNPJ}.xlsx` e `nfses_recebidas_{CNPJ}.xlsx`, 38 colunas,
   no mesmo layout que o robo antigo ja produz.

2. Planilha mensal consolidada, uma aba por competencia, uma linha por
   empresa. Essa e ATUALIZADA, nao reescrita do zero: a coluna IMPORTADO e
   preenchida a mao pelo Claudio depois de importar no Dominio, e seria
   apagada se o arquivo fosse recriado a cada rodada.
"""

from __future__ import annotations

import logging
import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

COLS_MENSAL = ["N°", "RAZÃO SOCIAL", "CNPJ", "XML - PRESTADOS",
               "XML - TOMADOS", "OBSERVAÇÃO", "IMPORTADO", "VALIDADE CERT"]

# Situacao de cada fluxo (prestados / tomados). Sao INDEPENDENTES: ter nota
# emitida e nao ter recebida (ou o contrario) e normal, nao e problema.
SIT_OK = "OK"
SIT_SEM_MOV = "SEM MOV"   # rodou certo, a empresa nao teve nota nesse fluxo
SIT_ERRO = "ERRO"         # nao deu para saber: certificado, TLS, etc.


def situacao_fluxo(status_empresa: str, quantidade: int) -> str:
    """Nao ter nota NAO e erro. So e ERRO quando a consulta nem chegou a rodar."""
    if status_empresa not in ("OK", "PARCIAL"):
        return SIT_ERRO
    return SIT_OK if quantidade > 0 else SIT_SEM_MOV


def _abrir(caminho: str):
    """openpyxl padrao; cai para read_only quando o arquivo veio com locale pt-BR."""
    try:
        return openpyxl.load_workbook(caminho)
    except TypeError:
        return openpyxl.load_workbook(caminho, read_only=False, data_only=False,
                                      rich_text=False)


def formatar_cnpj(cnpj: str) -> str:
    d = "".join(c for c in (cnpj or "") if c.isdigit())
    if len(d) != 14:
        return cnpj or ""
    return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"


# --------------------------------------------------------------------------
# 1. Planilha de notas
# --------------------------------------------------------------------------

def _notas_iguais(caminho: str, colunas: list[str], linhas: list[list]) -> bool:
    """
    A planilha que ja esta na pasta tem exatamente estas notas?

    Compara pelo CONTEUDO, nao pelos bytes: todo .xlsx guarda data/hora interna,
    entao dois arquivos com os mesmos dados nunca sao iguais byte a byte.
    """
    if not os.path.exists(caminho):
        return False
    try:
        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
        sh = wb[wb.sheetnames[0]]
        atual = list(sh.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return False

    esperado = [tuple(colunas)] + [tuple(l) for l in linhas]
    if len(atual) != len(esperado):
        return False

    def norm(v):
        if v is None:
            return ""
        if isinstance(v, float) and v == int(v):
            return str(int(v))
        return str(v).strip()

    return all(tuple(map(norm, a)) == tuple(map(norm, e))
               for a, e in zip(atual, esperado))


def gravar_notas(caminho: str, colunas: list[str], linhas: list[list],
                 nome_aba: str) -> str:
    """
    Escreve a planilha de notas. Grava mesmo sem nenhuma linha — uma planilha
    vazia com cabecalho e a prova de que a competencia foi processada e nao
    tinha nota, o que e diferente de nao ter rodado.
    """
    # Se a planilha ja existe com exatamente as mesmas notas, nao reescreve.
    # Arquivo ja presente na pasta nao e erro nem motivo para regravar.
    if _notas_iguais(caminho, colunas, linhas):
        logger.info("[PLANILHA] %s ja estava igual, mantido",
                    os.path.basename(caminho))
        return caminho

    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = nome_aba[:31]

    sh.append(colunas)
    for c in range(1, len(colunas) + 1):
        sh.cell(row=1, column=c).font = Font(bold=True)

    for linha in linhas:
        sh.append(linha)

    # largura só nas colunas que costumam ter texto longo
    larguras = {1: 52, 3: 19, 7: 38, 8: 26, 13: 20, 35: 42, 37: 42, 38: 20}
    for idx, larg in larguras.items():
        if idx <= len(colunas):
            sh.column_dimensions[get_column_letter(idx)].width = larg
    sh.freeze_panes = "A2"

    Path(caminho).parent.mkdir(parents=True, exist_ok=True)
    tmp = str(caminho) + ".tmp"
    wb.save(tmp)
    os.replace(tmp, caminho)
    logger.info("[PLANILHA] %s (%d nota(s))", os.path.basename(caminho), len(linhas))
    return caminho


# --------------------------------------------------------------------------
# 2. Planilha mensal consolidada
# --------------------------------------------------------------------------

def atualizar_mensal(caminho: str, competencia: str, resultados: list) -> str:
    """
    Cria/atualiza a aba da competencia, POR EMPRESA.

    Duas coisas sao preservadas do que ja estava na aba:

      - a coluna IMPORTADO, que o Claudio preenche a mao depois de importar
        no Dominio;
      - as LINHAS DE EMPRESAS QUE NAO ENTRARAM NESTA RODADA. Sem isso, rodar
        com -c ou --somente-erros apagaria da planilha todas as empresas que
        nao estavam na selecao.
    """
    caminho = str(caminho)
    if os.path.exists(caminho):
        wb = _abrir(caminho)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    aba = competencia
    anteriores: dict[str, list] = {}      # codigo -> linha inteira que ja estava la

    if aba in wb.sheetnames:
        antiga = wb[aba]
        cabec = [str(c.value or "") for c in antiga[1]]

        # TRAVA DE SEGURANCA. Em 25/08/2026 este robo reescreveu a aba 072026 da
        # planilha compartilhada e apagou os dados do OUTRO robo (o do ISS
        # Digital). A aba dele nao tem a coluna IMPORTADO, e tem uma coluna
        # "Tomados" extra. Se o cabecalho nao for o nosso, nao e nossa aba:
        # aborta em vez de destruir o trabalho de outro processo.
        if cabec and "IMPORTADO" not in cabec and any(c.strip() for c in cabec):
            raise PermissionError(
                f"A aba '{aba}' de '{os.path.basename(caminho)}' NAO foi criada por "
                f"este robo (nao tem a coluna IMPORTADO). Cabecalho encontrado: "
                f"{[c for c in cabec if c][:8]}. Recusando sobrescrever. "
                f"Aponte PLANILHA_MENSAL para um arquivo proprio do robo."
            )

        i_cod = cabec.index("N°") if "N°" in cabec else 0
        for linha in antiga.iter_rows(min_row=2, values_only=True):
            if not linha or len(linha) <= i_cod or linha[i_cod] in (None, ""):
                continue
            valores = ["" if v is None else v for v in linha]
            valores += [""] * (len(COLS_MENSAL) - len(valores))
            anteriores[str(linha[i_cod]).strip()] = valores[:len(COLS_MENSAL)]
        wb.remove(antiga)

    i_imp = COLS_MENSAL.index("IMPORTADO")

    # linhas desta rodada, sobrescrevendo o que houver da mesma empresa
    linhas_por_codigo = dict(anteriores)
    atualizadas = 0
    for r in resultados:
        cod = r.cliente.codigo_dominio
        vence = getattr(r, "cert_vence", None)
        antes = anteriores.get(cod)
        linhas_por_codigo[cod] = [
            cod,
            r.cliente.apelido or r.cliente.razao_social,
            formatar_cnpj(r.cliente.cnpj),
            situacao_fluxo(r.status, r.emitidas + r.emitidas_canceladas),
            situacao_fluxo(r.status, r.recebidas + r.recebidas_canceladas),
            r.observacao(),
            antes[i_imp] if antes else "",          # IMPORTADO e do Claudio
            vence.strftime("%d/%m/%Y") if vence else "",
        ]
        atualizadas += 1

    mantidas = len(linhas_por_codigo) - atualizadas
    logger.info("[MENSAL] aba %s: %d empresa(s) atualizada(s), %d mantida(s) "
                "de rodadas anteriores", aba, atualizadas, max(0, mantidas))
    marcados = sum(1 for l in anteriores.values() if str(l[i_imp]).strip())
    if marcados:
        logger.info("[MENSAL] %d marcacao(oes) de IMPORTADO preservada(s)", marcados)

    sh = wb.create_sheet(aba)
    sh.append(COLS_MENSAL)
    for c in range(1, len(COLS_MENSAL) + 1):
        sh.cell(row=1, column=c).font = Font(bold=True)

    def ordem(cod):
        return (0, int(cod)) if cod.isdigit() else (1, 0)

    for cod in sorted(linhas_por_codigo, key=ordem):
        sh.append(linhas_por_codigo[cod])

    for idx, larg in {1: 8, 2: 34, 3: 21, 4: 17, 5: 16, 6: 62, 7: 13, 8: 15}.items():
        sh.column_dimensions[get_column_letter(idx)].width = larg
    sh.freeze_panes = "A2"

    # abas em ordem decrescente de competencia (a mais nova primeiro)
    ordem = sorted(wb.sheetnames, key=lambda n: (n[2:] + n[:2]) if len(n) == 6 and n.isdigit() else n,
                   reverse=True)
    wb._sheets = [wb[n] for n in ordem]

    Path(caminho).parent.mkdir(parents=True, exist_ok=True)
    tmp = caminho + ".tmp"
    wb.save(tmp)
    os.replace(tmp, caminho)
    logger.info("[MENSAL] aba %s atualizada em %s", aba, os.path.basename(caminho))
    return caminho
